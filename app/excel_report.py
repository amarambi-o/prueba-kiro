"""
excel_report.py
Bank Modernization — Paso 9: Reporte Ejecutivo en Excel

Genera reports/executive_report.xlsx con hojas formateadas:
  1. Resumen Ejecutivo
  2. Scores & Dashboard
  3. Costos On-Prem vs AWS
  4. Equipo & Comparativa
  5. Roadmap & Riesgos

Uso:
    python excel_report.py --bucket <bucket> [--prefix bankdemo]
"""
import argparse, json, os
from datetime import datetime, timezone
import boto3
import aws_client
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

DEFAULT_PREFIX = "bankdemo"
REPORTS_DIR    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports")

# Paleta de colores
C_DARK_BLUE  = "1B2A4A"
C_MID_BLUE   = "2E5FA3"
C_LIGHT_BLUE = "D6E4F7"
C_GREEN      = "1E7E34"
C_GREEN_LIGHT= "D4EDDA"
C_YELLOW     = "856404"
C_YELLOW_BG  = "FFF3CD"
C_RED        = "721C24"
C_RED_BG     = "F8D7DA"
C_GRAY       = "6C757D"
C_GRAY_LIGHT = "F8F9FA"
C_WHITE      = "FFFFFF"
C_ORANGE     = "E65100"
C_ORANGE_BG  = "FFF3E0"

def _s3():
    return aws_client.s3()

def _get_json(bucket, key):
    print(f"  [GET] {key}")
    return json.loads(_s3().get_object(Bucket=bucket, Key=key)["Body"].read())

def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _cell_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _header_row(ws, row, values, bg=C_DARK_BLUE, fg=C_WHITE, size=10):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=size, bold=True, color=fg)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()

def _data_row(ws, row, values, bg=C_WHITE, bold=False, align="left", number_format=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=10, bold=bold, color="000000")
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = _border()
        if number_format and isinstance(val, (int, float)):
            c.number_format = number_format

def _title(ws, row, text, span, bg=C_DARK_BLUE, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=C_WHITE)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def _subtitle(ws, row, text, span, bg=C_MID_BLUE, size=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=C_WHITE)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def _score_color(score, inverted=False):
    v = (100 - score) if inverted else score
    if v >= 70: return C_RED_BG, C_RED
    if v >= 40: return C_YELLOW_BG, C_YELLOW
    return C_GREEN_LIGHT, C_GREEN

def _score_label(score, inverted=False):
    v = (100 - score) if inverted else score
    if v >= 70: return "🔴 Alto Riesgo"
    if v >= 40: return "🟡 Moderado"
    return "🟢 Adecuado"


# ---------------------------------------------------------------------------
# Hoja 1 — Resumen Ejecutivo
# ---------------------------------------------------------------------------
def _sheet_resumen(wb, es, fin, scores, fecha):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([3,22,18,18,18,18], 1):
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 10

    _title(ws, 2, "BANK MODERNIZATION READINESS ADVISOR", 6, C_DARK_BLUE, 14)
    _title(ws, 3, f"Reporte Ejecutivo — payments-core / BankDemo  |  {fecha}  |  Confidencial", 6, C_MID_BLUE, 10)

    ws.row_dimensions[4].height = 12

    _subtitle(ws, 5, "RESUMEN EJECUTIVO", 6)
    ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=6)
    c = ws.cell(row=6, column=1, value=es["headline"])
    c.font = Font(name="Calibri", size=10, color="000000")
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    c.fill = _fill(C_LIGHT_BLUE)
    ws.row_dimensions[6].height = 50

    ws.row_dimensions[9].height = 12
    _subtitle(ws, 10, "INDICADORES CLAVE", 6)
    _header_row(ws, 11, ["", "Indicador", "Valor", "Unidad", "Tendencia", ""], C_MID_BLUE)

    kpis = [
        ("Estrategia recomendada", es["recommended_strategy"].upper(), "", "✅ Definida"),
        ("Duración implementación", es["effort_weeks"], "semanas", "📅 20 semanas"),
        ("Inversión total", fin["total_investment_usd"], "USD", "💰 Incluye contingencia 15%"),
        ("Ahorro neto anual", fin["annual_savings_usd"], "USD/año", "📈 Post go-live"),
        ("Payback", fin["payback_months"], "meses", "⏱ ~2 años"),
        ("ROI a 3 años", fin["roi_3_years_pct"], "%", "📊 Positivo"),
        ("Ahorro acumulado 3 años", fin["cumulative_savings_3y_usd"], "USD", "💵 Neto"),
        ("Hallazgos de compliance", 130, "findings", "⚠ Requieren remediación"),
        ("Riesgo regulatorio", scores["regulatory_risk_score"], "/ 100", "🟡 Medio"),
        ("DQ Score", scores["data_quality_score"], "/ 100", "🟡 Moderado"),
    ]
    for i, (ind, val, unit, tend) in enumerate(kpis):
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, 12+i, ["", ind, val, unit, tend, ""], bg)
        ws.cell(12+i, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(12+i, 4).alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Hoja 2 — Scores & Dashboard
# ---------------------------------------------------------------------------
def _sheet_scores(wb, scores, es):
    ws = wb.create_sheet("Scores & Dashboard")
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([3,30,14,20,30,3], 1):
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 10

    _title(ws, 2, "DASHBOARD DE SCORES — payments-core", 6, C_DARK_BLUE, 13)
    ws.row_dimensions[3].height = 8

    _subtitle(ws, 4, "SCORES DE READINESS Y RIESGO", 6)
    _header_row(ws, 5, ["", "Dimensión", "Score", "Estado", "Interpretación", ""], C_MID_BLUE)

    score_rows = [
        ("Cloud Readiness",      scores["cloud_readiness_score"],      True,  "Preparación para migrar a nube"),
        ("Data Quality",         scores["data_quality_score"],         True,  "Calidad de los datos del sistema"),
        ("Security Risk",        scores["security_risk_score"],        False, "Nivel de riesgo de seguridad"),
        ("Compliance Risk",      scores["compliance_risk_score"],      False, "Exposición regulatoria"),
        ("PCI-DSS Readiness",    scores["pci_readiness_score"],        True,  "Preparación para PCI-DSS v4.0"),
        ("SOX Traceability",     scores["sox_traceability_score"],     True,  "Trazabilidad de transacciones"),
        ("PII Exposure",         scores["pii_exposure_score"],         False, "Exposición de datos personales"),
        ("Encryption Coverage",  scores["encryption_coverage_score"],  True,  "Cobertura de cifrado en reposo"),
        ("Auditability",         scores["auditability_score"],         True,  "Capacidad de auditoría"),
        ("Regulatory Risk",      scores["regulatory_risk_score"],      False, "Riesgo regulatorio compuesto"),
        ("Migration Complexity", es["migration_complexity_score"],     False, "Complejidad de la migración"),
    ]
    for i, (dim, score, inv, interp) in enumerate(score_rows):
        bg_cell, _ = _score_color(score, inv)
        bg_row = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        label = _score_label(score, inv)
        _data_row(ws, 6+i, ["", dim, f"{score} / 100", label, interp, ""], bg_row)
        ws.cell(6+i, 3).fill = _fill(bg_cell)
        ws.cell(6+i, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(6+i, 3).font = Font(name="Calibri", size=10, bold=True, color="000000")
        ws.cell(6+i, 4).fill = _fill(bg_cell)
        ws.cell(6+i, 4).alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[17].height = 12
    _subtitle(ws, 18, "COMPLEJIDAD DE MIGRACIÓN POR DIMENSIÓN", 6)
    _header_row(ws, 19, ["", "Dimensión", "Nivel", "Detalle", "", ""], C_MID_BLUE)
    complexity = [
        ("Aplicación",   "HIGH",   "Monolito Python sin microservicios"),
        ("Datos",        "MEDIUM", f"SQL Server on-prem · DQ Score {scores['data_quality_score']}/100"),
        ("Seguridad",    "HIGH",   "Sin auth, credenciales hardcodeadas, sin cifrado"),
        ("Compliance",   "MEDIUM", "3 frameworks en riesgo activo"),
        ("Operaciones",  "MEDIUM", "Sin SLA, sin monitoreo, sin HA"),
    ]
    for i, (dim, nivel, det) in enumerate(complexity):
        bg = C_RED_BG if nivel == "HIGH" else C_YELLOW_BG
        bg_row = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, 20+i, ["", dim, nivel, det, "", ""], bg_row)
        ws.cell(20+i, 3).fill = _fill(bg)
        ws.cell(20+i, 3).alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Hoja 3 — Costos On-Prem vs AWS
# ---------------------------------------------------------------------------
def _sheet_costos(wb, fin):
    ws = wb.create_sheet("Costos On-Prem vs AWS")
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([3,32,16,16,16,3], 1):
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 10

    _title(ws, 2, "ANÁLISIS DE COSTOS — ON-PREMISES vs AWS", 6, C_DARK_BLUE, 13)
    ws.row_dimensions[3].height = 8

    _subtitle(ws, 4, "COSTO ANUAL ACTUAL vs COSTO ANUAL POST-MIGRACIÓN", 6)
    _header_row(ws, 5, ["", "Concepto", "On-Premises (actual)", "AWS (post-migración)", "Ahorro anual", ""], C_MID_BLUE)

    cost_rows = [
        ("Licencias SQL Server Enterprise",  85000,   0,      85000),
        ("Hardware e infraestructura",        48000,   0,      48000),
        ("Operaciones IT (FTEs infra)",       72000,   0,      72000),
        ("Auditorías manuales de compliance", 45500,   0,      45500),
        ("Incidentes de seguridad",           42000,   5000,   37000),
        ("Infraestructura cloud (AWS)",       0,       fin["aws_annual_cost_usd"], -fin["aws_annual_cost_usd"]),
    ]
    fmt = '#,##0 "USD"'
    for i, (concepto, onprem, aws, ahorro) in enumerate(cost_rows):
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, 6+i, ["", concepto, onprem, aws, ahorro, ""], bg, number_format=fmt)
        for col in [3,4,5]:
            ws.cell(6+i, col).alignment = Alignment(horizontal="right", vertical="center")
        if ahorro > 0:
            ws.cell(6+i, 5).fill = _fill(C_GREEN_LIGHT)
            ws.cell(6+i, 5).font = Font(name="Calibri", size=10, bold=True, color=C_GREEN)
        elif ahorro < 0:
            ws.cell(6+i, 5).fill = _fill(C_RED_BG)
            ws.cell(6+i, 5).font = Font(name="Calibri", size=10, color=C_RED)

    total_row = 12
    _data_row(ws, total_row, ["", "TOTAL", fin["onpremises_annual_cost_usd"],
                               fin["aws_annual_cost_usd"]+5000, fin["annual_savings_usd"], ""],
              C_LIGHT_BLUE, bold=True, number_format=fmt)
    for col in [3,4,5]:
        ws.cell(total_row, col).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(total_row, 5).fill = _fill(C_GREEN_LIGHT)
    ws.cell(total_row, 5).font = Font(name="Calibri", size=10, bold=True, color=C_GREEN)

    ws.row_dimensions[13].height = 12
    _subtitle(ws, 14, "RESUMEN FINANCIERO DE LA INVERSIÓN", 6)
    _header_row(ws, 15, ["", "Concepto", "Monto", "Notas", "", ""], C_MID_BLUE)

    inv_rows = [
        ("Costo implementación (equipo)",  fin["implementation_cost_usd"],  "5 personas · 20 semanas"),
        ("Contingencia (15%)",             fin["contingency_usd"],           "Buffer de riesgo estándar"),
        ("Inversión total",                fin["total_investment_usd"],      "One-time"),
        ("AWS mensual post go-live",       fin["aws_monthly_cost_usd"],      "Recurrente"),
        ("AWS anual post go-live",         fin["aws_annual_cost_usd"],       "Recurrente"),
        ("Ahorro neto anual",              fin["annual_savings_usd"],        "A partir del go-live"),
        ("Payback",                        fin["payback_months"],            "meses"),
        ("ROI a 3 años",                   fin["roi_3_years_pct"],           "%"),
        ("Ahorro acumulado neto 3 años",   fin["cumulative_savings_3y_usd"], "USD"),
    ]
    for i, (concepto, monto, nota) in enumerate(inv_rows):
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        bold = concepto in ("Inversión total", "Ahorro neto anual", "ROI a 3 años")
        _data_row(ws, 16+i, ["", concepto, monto, nota, "", ""], bg, bold=bold, number_format=fmt)
        ws.cell(16+i, 3).alignment = Alignment(horizontal="right", vertical="center")
        if "Ahorro" in concepto or "ROI" in concepto:
            ws.cell(16+i, 3).fill = _fill(C_GREEN_LIGHT)
            ws.cell(16+i, 3).font = Font(name="Calibri", size=10, bold=bold, color=C_GREEN)


# ---------------------------------------------------------------------------
# Hoja 4 — Equipo & Comparativa
# ---------------------------------------------------------------------------
def _sheet_equipo(wb, team, fin, weeks):
    ws = wb.create_sheet("Equipo & Comparativa")
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([3,28,14,12,14,16,3], 1):
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 10

    _title(ws, 2, "EQUIPO REQUERIDO Y COMPARATIVA CON MÉTODO TRADICIONAL", 7, C_DARK_BLUE, 13)
    ws.row_dimensions[3].height = 8

    _subtitle(ws, 4, "FASE 1 — ASSESSMENT: TRADICIONAL vs KIRO + AWS", 7)
    _header_row(ws, 5, ["", "Dimensión", "Consultoría Tradicional", "Kiro + AWS MCP", "Ventaja", "", ""], C_MID_BLUE)
    assess = [
        ("Duración",             "6 – 8 semanas",          "< 1 hora (pipeline)",    "99% más rápido"),
        ("Personas requeridas",  "6 – 8 consultores",      "0 (automatizado)",       "100% automatizado"),
        ("Costo",                "USD 400K – 800K",        "Incluido en plataforma", "Ahorro inmediato"),
        ("Cobertura de datos",   "Muestra (~10%)",         "100% de registros",      "Cobertura total"),
        ("Reportes",             "Word/Excel estáticos",   "Markdown trazable",      "Regenerable en cada run"),
        ("Trazabilidad",         "Manual, sin fuente",     "JSON de origen por hallazgo", "Auditable"),
    ]
    for i, (dim, trad, kiro, vent) in enumerate(assess):
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, 6+i, ["", dim, trad, kiro, vent, "", ""], bg)
        ws.cell(6+i, 4).fill = _fill(C_GREEN_LIGHT)
        ws.cell(6+i, 4).font = Font(name="Calibri", size=10, color=C_GREEN, bold=True)

    ws.row_dimensions[12].height = 12
    _subtitle(ws, 13, "FASE 2 — IMPLEMENTACIÓN CON KIRO + AWS (20 SEMANAS)", 7)
    _header_row(ws, 14, ["", "Rol", "Dedicación", "Personas", "Duración", "Costo estimado", ""], C_MID_BLUE)

    rates = {"Cloud Architect": 200, "DevOps Engineer": 150, "Data Engineer": 160,
             "Security Engineer": 170, "Project Manager": 120}
    fmt = '#,##0 "USD"'
    total_cost = 0
    for i, m in enumerate(team):
        role = m["role"]; ded = m["dedication"]; hc = m["headcount"]
        ded_pct = int(ded.replace("%","")) / 100
        cost = round(weeks * 40 * ded_pct * rates.get(role,150) * hc / 1000) * 1000
        total_cost += cost
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, 15+i, ["", role, ded, hc, f"{weeks} semanas", cost, ""], bg, number_format=fmt)
        ws.cell(15+i, 6).alignment = Alignment(horizontal="right", vertical="center")

    total_r = 15 + len(team)
    _data_row(ws, total_r, ["", "TOTAL EQUIPO", "", sum(m["headcount"] for m in team),
                             f"{weeks} semanas", fin["implementation_cost_usd"], ""],
              C_LIGHT_BLUE, bold=True, number_format=fmt)
    ws.cell(total_r, 6).alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[total_r+1].height = 12
    _subtitle(ws, total_r+2, "COMPARATIVA DE IMPLEMENTACIÓN: TRADICIONAL vs KIRO + AWS", 7)
    _header_row(ws, total_r+3, ["", "Dimensión", "Consultoría Tradicional", "Kiro + AWS", "Ventaja", "", ""], C_MID_BLUE)
    impl = [
        ("Duración",             "18 – 24 meses",                    "20 semanas (~5 meses)",                  "4x más rápido"),
        ("Personas requeridas",  "12 – 15",                          "5",                                      "67% menos"),
        ("Costo implementación", "USD 1,200,000 – 2,000,000",        f"USD {fin['total_investment_usd']:,}",   "Hasta USD 1,434,000 menos"),
        ("Reportes de avance",   "Manuales, por hitos",              "Automatizados en cada ejecución",        "Visibilidad continua"),
        ("Compliance integrado", "Post-implementación",              "Desde el día 1 (pipeline)",              "Riesgo reducido desde inicio"),
    ]
    for i, (dim, trad, kiro, vent) in enumerate(impl):
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        _data_row(ws, total_r+4+i, ["", dim, trad, kiro, vent, "", ""], bg)
        ws.cell(total_r+4+i, 4).fill = _fill(C_GREEN_LIGHT)
        ws.cell(total_r+4+i, 4).font = Font(name="Calibri", size=10, color=C_GREEN, bold=True)


# ---------------------------------------------------------------------------
# Hoja 5 — Roadmap & Riesgos
# ---------------------------------------------------------------------------
def _sheet_roadmap(wb, phases, risks, strategy):
    ws = wb.create_sheet("Roadmap & Riesgos")
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([3,20,12,40,3], 1):
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 10

    _title(ws, 2, "ROADMAP DE IMPLEMENTACIÓN Y GESTIÓN DE RIESGOS", 5, C_DARK_BLUE, 13)
    ws.row_dimensions[3].height = 8

    _subtitle(ws, 4, "HITOS CLAVE — ESTRATEGIA HYBRID (20 SEMANAS)", 5)
    _header_row(ws, 5, ["", "Hito", "Semana", "Descripción", ""], C_MID_BLUE)
    phase_colors = [C_LIGHT_BLUE, "E8F5E9", "FFF9C4", "FCE4EC"]
    for i, m in enumerate(phases):
        bg = phase_colors[i % len(phase_colors)]
        _data_row(ws, 6+i, ["", m["milestone"], f"Semana {m['week']}", "", ""], bg)
        ws.cell(6+i, 3).alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[6+len(phases)].height = 12
    r = 7 + len(phases)
    _subtitle(ws, r, "FASES DE MIGRACIÓN", 5)
    _header_row(ws, r+1, ["", "Fase", "Semanas", "Acciones principales", ""], C_MID_BLUE)
    migration_phases = [
        ("1 — Fundamentos de Seguridad", "1 – 4",  "CloudTrail + KMS + Secrets Manager · Security Hub + AWS Config"),
        ("2 — Gobernanza",               "5 – 8",  "Control Tower · Audit Manager · Glue + Lake Formation"),
        ("3 — Migración Híbrida",        "9 – 16", "Rehost EC2/RDS · Replatform Aurora · Refactor módulos en EKS"),
        ("4 — Consolidación",            "17 – 20","Unificación arquitectura · Auditoría PCI-DSS · Documentación SOX"),
    ]
    for i, (fase, sems, acciones) in enumerate(migration_phases):
        bg = phase_colors[i]
        _data_row(ws, r+2+i, ["", fase, sems, acciones, ""], bg)
        ws.cell(r+2+i, 3).alignment = Alignment(horizontal="center", vertical="center")

    r2 = r + 2 + len(migration_phases) + 1
    ws.row_dimensions[r2-1].height = 12
    _subtitle(ws, r2, "GESTIÓN DE RIESGOS", 5)
    _header_row(ws, r2+1, ["", "Riesgo", "Probabilidad", "Impacto · Mitigación", ""], C_MID_BLUE)
    for i, risk in enumerate(risks):
        prob_bg = C_RED_BG if risk["probability"] == "HIGH" else C_YELLOW_BG
        bg = C_GRAY_LIGHT if i % 2 == 0 else C_WHITE
        detail = f"{risk['impact']} — {risk['mitigation']}"
        _data_row(ws, r2+2+i, ["", risk["risk"], risk["probability"], detail, ""], bg)
        ws.cell(r2+2+i, 3).fill = _fill(prob_bg)
        ws.cell(r2+2+i, 3).alignment = Alignment(horizontal="center", vertical="center")

    r3 = r2 + 2 + len(risks) + 1
    ws.row_dimensions[r3-1].height = 12
    _subtitle(ws, r3, "RIESGO DE NO ACTUAR", 5)
    _header_row(ws, r3+1, ["", "Concepto", "Estimación", "", ""], C_RED, C_WHITE)
    inaction = [
        ("Multas regulatorias PCI-DSS / SOX", strategy["regulatory_risk"]["estimated_fine_range_usd"]),
        ("Exposición GDPR",                   "Hasta 4% de facturación global"),
        ("Crecimiento deuda técnica",          "~25% anual"),
        ("Costo total de inacción a 3 años",   "USD 936,000 – 4,236,000"),
    ]
    for i, (concepto, est) in enumerate(inaction):
        bg = C_RED_BG if i % 2 == 0 else C_WHITE
        _data_row(ws, r3+2+i, ["", concepto, est, "", ""], bg)
        ws.cell(r3+2+i, 2).font = Font(name="Calibri", size=10, color=C_RED)


# ---------------------------------------------------------------------------
# Orquestador principal
# ---------------------------------------------------------------------------
def generate(bucket, prefix):
    summary    = _get_json(bucket, f"{prefix}/output/modernization/modernization_summary.json")
    estimation = _get_json(bucket, f"{prefix}/output/modernization/project_estimation.json")
    strategy   = _get_json(bucket, f"{prefix}/output/modernization/migration_strategy.json")

    es     = summary["executive_summary"]
    scores = summary["input_scores"]
    fin    = estimation["financials"]
    team   = estimation["team"]
    phases = estimation["milestones"]
    risks  = estimation["risks"]
    weeks  = estimation["project_duration_weeks"]
    fecha  = datetime.now(timezone.utc).strftime("%B %Y")

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja default

    _sheet_resumen(wb, es, fin, scores, fecha)
    _sheet_scores(wb, scores, es)
    _sheet_costos(wb, fin)
    _sheet_equipo(wb, team, fin, weeks)
    _sheet_roadmap(wb, phases, risks, strategy)

    os.makedirs(REPORTS_DIR, exist_ok=True)
    out_path = os.path.join(REPORTS_DIR, "executive_report.xlsx")
    wb.save(out_path)
    print(f"  → {out_path}")

    with open(out_path, "rb") as f:
        _s3().put_object(
            Bucket=bucket,
            Key=f"{prefix}/output/executive_report.xlsx",
            Body=f.read(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    print(f"  → s3://{bucket}/{prefix}/output/executive_report.xlsx")
    return out_path


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--bucket", default=os.environ.get("S3_BUCKET", ""), required=not os.environ.get("S3_BUCKET"))
    p.add_argument("--prefix", default=os.environ.get("S3_PREFIX", DEFAULT_PREFIX))
    a = p.parse_args()
    generate(a.bucket, a.prefix)


if __name__ == "__main__":
    main()
