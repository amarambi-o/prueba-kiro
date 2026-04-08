"""
generate_excel_report.py — Reporte Ejecutivo Excel
Lee JSONs de S3 y genera reports/executive_report.xlsx con 5 hojas.
Uso: python app/generators/generate_excel_report.py
"""
import argparse, io, json, os
from datetime import datetime, timezone

import boto3

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("pip install openpyxl")

BUCKET  = "bank-modernization-kiro"
PREFIX  = "bankdemo"
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "reports")

C_DARK_BLUE   = "0F2D5E"
C_BLUE        = "0F62FE"
C_LIGHT_BLUE  = "D0E4FF"
C_GREEN       = "198038"
C_LIGHT_GREEN = "DEFBE6"
C_RED         = "DA1E28"
C_LIGHT_RED   = "FFD7D9"
C_ORANGE      = "FF832B"
C_LIGHT_ORANGE= "FFE8CC"
C_YELLOW      = "F1C21B"
C_GRAY        = "F4F4F4"
C_WHITE       = "FFFFFF"
C_DARK_GRAY   = "525252"


def _score_color(score, inverted=False):
    v = (100 - score) if inverted else score
    if v >= 70: return C_LIGHT_GREEN
    if v >= 40: return C_LIGHT_ORANGE
    return C_LIGHT_RED

def _score_label(score, inverted=False):
    v = (100 - score) if inverted else score
    if v >= 70: return "🟢 Adecuado"
    if v >= 40: return "🟡 Moderado"
    return "🔴 Alto Riesgo"

# ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────────────

def _hoja_resumen(wb, r, c, m, dq):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [28, 22, 18, 22, 22, 22])

    es = m.get("executive_summary", {})
    strategy  = es.get("recommended_strategy", "N/A").upper()
    effort    = es.get("effort_weeks", "N/A")
    inv       = es.get("total_investment_usd", 0)
    savings   = es.get("annual_savings_usd", 0)
    payback   = es.get("payback_months", 0)
    roi       = es.get("roi_3_years_pct", 0)
    cum3y     = es.get("cumulative_savings_3y_usd", 0) if "cumulative_savings_3y_usd" in es else round(savings * 3 - inv)
    findings  = m.get("compliance_findings_analyzed", 0)
    reg_risk  = c.get("regulatory_risk_score", 0)
    dq_score  = r.get("data_quality_score", 0)
    date_str  = datetime.now(timezone.utc).strftime("%B %Y")

    ws.row_dimensions[1].height = 14
    _title(ws, 2, "BANK MODERNIZATION READINESS ADVISOR", C_DARK_BLUE, 16)
    ws.row_dimensions[2].height = 32
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    sub = ws.cell(row=3, column=1, value=f"Reporte Ejecutivo — payments-core / BankDemo  |  {date_str}  |  Confidencial")
    sub.fill = _fill(C_BLUE); sub.font = _font(size=11, color=C_WHITE, italic=True)
    sub.alignment = _align("center")
    ws.row_dimensions[3].height = 20

    ws.row_dimensions[5].height = 14
    _section(ws, 6, "RESUMEN EJECUTIVO")
    ws.row_dimensions[6].height = 22
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)
    headline = m.get("executive_summary", {}).get("headline", "")
    h = ws.cell(row=7, column=1, value=headline)
    h.fill = _fill(C_LIGHT_BLUE); h.font = _font(size=10, italic=True)
    h.alignment = _align("left", wrap=True)
    ws.row_dimensions[7].height = 48

    ws.row_dimensions[9].height = 14
    _section(ws, 10, "INDICADORES CLAVE")
    ws.row_dimensions[10].height = 22
    _header_row(ws, 11, ["Indicador", "Valor", "Unidad", "Tendencia", "", ""])
    kpis = [
        ("Estrategia recomendada",    strategy,          "",          "✅ Definida"),
        ("Duración implementación",   effort,            "semanas",   f"📅 {effort} semanas"),
        ("Inversión total",           f"{inv:,}",        "USD",       "💰 Incluye contingencia 15%"),
        ("Ahorro neto anual",         f"{savings:,}",    "USD/año",   "📈 Post go-live"),
        ("Payback",                   payback,           "meses",     "⏱ ~2 años"),
        ("ROI a 3 años",              roi,               "%",         "📊 Positivo"),
        ("Ahorro acumulado 3 años",   f"{cum3y:,}",      "USD",       "💵 Neto"),
        ("Hallazgos de compliance",   findings,          "findings",  "⚠ Requieren remediación"),
        ("Riesgo regulatorio",        f"{reg_risk} / 100","",         "🟡 Medio" if 40 <= reg_risk < 70 else "🔴 Alto"),
        ("DQ Score",                  f"{dq_score} / 100","",         _score_label(dq_score)),
    ]
    for i, (ind, val, unit, tend) in enumerate(kpis):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        _data_row(ws, 12 + i, [ind, val, unit, tend, "", ""], bg)

# ── Hoja 2: Scores & Dashboard ────────────────────────────────────────────────

def _hoja_scores(wb, r, c, m):
    ws = wb.create_sheet("Scores & Dashboard")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [30, 18, 20, 36, 18, 18])

    _title(ws, 1, "DASHBOARD DE SCORES — payments-core", C_DARK_BLUE, 14)
    ws.row_dimensions[1].height = 28

    _section(ws, 3, "SCORES DE READINESS Y RIESGO")
    _header_row(ws, 4, ["Dimension", "Score", "Estado", "Interpretacion", "", ""])

    scores = [
        ("Cloud Readiness",    r.get("cloud_readiness_score", 0),    False, "Preparacion para migrar a nube"),
        ("Data Quality",       r.get("data_quality_score", 0),       False, "Calidad de los datos del sistema"),
        ("Security Risk",      r.get("security_risk_score", 0),      True,  "Nivel de riesgo de seguridad"),
        ("Compliance Risk",    r.get("compliance_risk_score", 0),    True,  "Exposicion regulatoria"),
        ("PCI-DSS Readiness",  c.get("pci_readiness_score", 0),      False, "Preparacion para PCI-DSS v4.0"),
        ("SOX Traceability",   c.get("sox_traceability_score", 0),   False, "Trazabilidad de transacciones"),
        ("PII Exposure",       c.get("pii_exposure_score", 0),       True,  "Exposicion de datos personales"),
        ("Encryption Coverage",c.get("encryption_coverage_score", 0),False, "Cobertura de cifrado en reposo"),
        ("Auditability",       c.get("auditability_score", 0),       False, "Capacidad de auditoria"),
        ("Regulatory Risk",    c.get("regulatory_risk_score", 0),    True,  "Riesgo regulatorio compuesto"),
        ("Migration Complexity",m.get("executive_summary", {}).get("migration_complexity_score", 0), True, "Complejidad de la migracion"),
    ]
    for i, (dim, score, inv, interp) in enumerate(scores):
        bg = _score_color(score, inv)
        label = _score_label(score, inv)
        _data_row(ws, 5 + i, [dim, f"{score} / 100", label, interp, "", ""], bg)

    dq = r.get("data_quality_score", 0)
    row = 5 + len(scores) + 2
    _section(ws, row, "COMPLEJIDAD DE MIGRACION POR DIMENSION")
    _header_row(ws, row + 1, ["Dimension", "Nivel", "Detalle", "", "", ""])
    complexity = [
        ("Aplicacion",  "HIGH",   "Monolito Python sin microservicios"),
        ("Datos",       "MEDIUM" if dq >= 70 else "HIGH", f"SQL Server on-prem - DQ Score {dq}/100"),
        ("Seguridad",   "HIGH",   "Sin auth, credenciales hardcodeadas, sin cifrado"),
        ("Compliance",  "MEDIUM", "3 frameworks en riesgo activo"),
        ("Operaciones", "MEDIUM", "Sin SLA, sin monitoreo, sin HA"),
    ]
    for i, (dim, nivel, det) in enumerate(complexity):
        bg = C_LIGHT_RED if nivel == "HIGH" else C_LIGHT_ORANGE
        _data_row(ws, row + 2 + i, [dim, nivel, det, "", "", ""], bg)

# ── Hoja 3: Costos ────────────────────────────────────────────────────────────

def _hoja_costos(wb, bc, pe):
    ws = wb.create_sheet("Costos On-Prem vs AWS")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [36, 22, 22, 22, 18, 18])

    _title(ws, 1, "ANALISIS DE COSTOS - ON-PREMISES vs AWS", C_DARK_BLUE, 14)
    ws.row_dimensions[1].height = 28

    _section(ws, 3, "COSTO ANUAL ACTUAL vs COSTO ANUAL POST-MIGRACION")
    _header_row(ws, 4, ["Concepto", "On-Premises (actual)", "AWS (post-migracion)", "Ahorro anual", "", ""])

    fin = bc.get("financial_summary", {}) if bc else {}
    aws_annual = fin.get("aws_annual_cost_usd", 11400)

    cost_rows = [
        ("Licencias SQL Server Enterprise",  85000,  0,          85000),
        ("Hardware e infraestructura",        48000,  0,          48000),
        ("Operaciones IT (FTEs infra)",       72000,  0,          72000),
        ("Auditorias manuales de compliance", 45500,  0,          45500),
        ("Incidentes de seguridad",           42000,  5000,       37000),
        ("Infraestructura cloud (AWS)",       0,      aws_annual, -aws_annual),
    ]
    total_onprem = sum(x[1] for x in cost_rows)
    total_aws    = sum(x[2] for x in cost_rows)
    total_saving = total_onprem - total_aws

    for i, (concept, onprem, aws, saving) in enumerate(cost_rows):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        _data_row(ws, 5 + i, [concept, f"{onprem:,} USD", f"{aws:,} USD", f"{saving:,} USD", "", ""], bg)
    _data_row(ws, 5 + len(cost_rows), ["TOTAL", f"{total_onprem:,} USD", f"{total_aws:,} USD", f"{total_saving:,} USD", "", ""],
              C_LIGHT_BLUE, bold=True)

    row = 5 + len(cost_rows) + 3
    _section(ws, row, "RESUMEN FINANCIERO DE LA INVERSION")
    _header_row(ws, row + 1, ["Concepto", "Monto", "Notas", "", "", ""])

    inv  = fin.get("total_investment_usd", 0)
    impl = pe.get("financials", {}).get("implementation_cost_usd", 0) if pe else 0
    cont = pe.get("financials", {}).get("contingency_usd", 0) if pe else 0
    net  = fin.get("net_annual_savings_usd", total_saving)
    pb   = fin.get("payback_months", 0)
    roi  = fin.get("roi_3_years_pct", 0)
    npv  = fin.get("npv_3_years_usd", 0)

    fin_rows = [
        ("Costo implementacion (equipo)", f"{impl:,} USD",  "Equipo - semanas"),
        ("Contingencia (15%)",            f"{cont:,} USD",  "Buffer de riesgo estandar"),
        ("Inversion total",               f"{inv:,} USD",   "One-time"),
        ("AWS mensual post go-live",      f"{aws_annual//12:,} USD", "Recurrente"),
        ("AWS anual post go-live",        f"{aws_annual:,} USD",     "Recurrente"),
        ("Ahorro neto anual",             f"{net:,} USD",   "A partir del go-live"),
        ("Payback",                       f"{pb} meses",    ""),
        ("ROI a 3 anos",                  f"{roi}%",        ""),
        ("Ahorro acumulado neto 3 anos",  f"{npv:,} USD",   "USD"),
    ]
    for i, (concept, monto, nota) in enumerate(fin_rows):
        bg = C_LIGHT_GREEN if "Ahorro" in concept or "ROI" in concept else (C_GRAY if i % 2 == 0 else C_WHITE)
        _data_row(ws, row + 2 + i, [concept, monto, nota, "", "", ""], bg)

# ── Hoja 4: Equipo & Comparativa ─────────────────────────────────────────────

def _hoja_equipo(wb, pe, ms):
    ws = wb.create_sheet("Equipo & Comparativa")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [30, 22, 18, 18, 22, 22])

    effort = pe.get("project_duration_weeks", 20) if pe else 20
    _title(ws, 1, "EQUIPO REQUERIDO Y COMPARATIVA CON METODO TRADICIONAL", C_DARK_BLUE, 14)
    ws.row_dimensions[1].height = 28

    _section(ws, 3, "FASE 1 - ASSESSMENT: TRADICIONAL vs KIRO + AWS")
    _header_row(ws, 4, ["Dimension", "Consultoria Tradicional", "Kiro + AWS MCP", "Ventaja", "", ""])
    comp1 = [
        ("Duracion",             "6 - 8 semanas",          "< 1 hora (pipeline)",         "99% mas rapido"),
        ("Personas requeridas",  "6 - 8 consultores",      "0 (automatizado)",            "100% automatizado"),
        ("Costo",                "USD 400K - 800K",        "Incluido en plataforma",      "Ahorro inmediato"),
        ("Cobertura de datos",   "Muestra (~10%)",         "100% de registros",           "Cobertura total"),
        ("Reportes",             "Word/Excel estaticos",   "Markdown trazable",           "Regenerable en cada run"),
        ("Trazabilidad",         "Manual, sin fuente",     "JSON de origen por hallazgo", "Auditable"),
    ]
    for i, row_data in enumerate(comp1):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        _data_row(ws, 5 + i, list(row_data) + ["", ""], bg)

    team = pe.get("team", []) if pe else []
    row = 5 + len(comp1) + 2
    _section(ws, row, f"FASE 2 - IMPLEMENTACION CON KIRO + AWS ({effort} SEMANAS)")
    _header_row(ws, row + 1, ["Rol", "Dedicacion", "Personas", "Duracion", "Costo estimado", ""])
    for i, t in enumerate(team):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        _data_row(ws, row + 2 + i, [t.get("role",""), t.get("dedication",""), f"{t.get('headcount',1)} USD", f"{effort} semanas", "", ""], bg)
    fin = pe.get("financials", {}) if pe else {}
    impl = fin.get("implementation_cost_usd", 0)
    _data_row(ws, row + 2 + len(team), ["TOTAL EQUIPO", "", f"{len(team)} USD", f"{effort} semanas", f"{impl:,} USD", ""],
              C_LIGHT_BLUE, bold=True)

    row2 = row + 2 + len(team) + 3
    _section(ws, row2, "COMPARATIVA DE IMPLEMENTACION: TRADICIONAL vs KIRO + AWS")
    _header_row(ws, row2 + 1, ["Dimension", "Consultoria Tradicional", "Kiro + AWS", "Ventaja", "", ""])
    comp2 = [
        ("Duracion",               "18 - 24 meses",               f"{effort} semanas (~5 meses)", "4x mas rapido"),
        ("Personas requeridas",    "12 - 15",                     "5",                            "67% menos"),
        ("Costo implementacion",   "USD 1,200,000 - 2,000,000",   f"USD {impl:,}",                "Hasta USD 1,434,000 menos"),
        ("Reportes de avance",     "Manuales, por hitos",         "Automatizados en cada ejecucion","Visibilidad continua"),
        ("Compliance integrado",   "Post-implementacion",         "Desde el dia 1 (pipeline)",    "Riesgo reducido desde inicio"),
    ]
    for i, row_data in enumerate(comp2):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        _data_row(ws, row2 + 2 + i, list(row_data) + ["", ""], bg)


# ── Hoja 5: Roadmap & Riesgos ─────────────────────────────────────────────────

def _hoja_roadmap(wb, ms, pe):
    ws = wb.create_sheet("Roadmap & Riesgos")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [28, 14, 46, 18, 18, 18])

    strategy = ms.get("recommended_strategy", "N/A").upper() if ms else "N/A"
    effort   = pe.get("project_duration_weeks", 20) if pe else 20
    _title(ws, 1, f"ROADMAP DE IMPLEMENTACION Y GESTION DE RIESGOS", C_DARK_BLUE, 14)
    ws.row_dimensions[1].height = 28

    milestones = pe.get("milestones", []) if pe else []
    _section(ws, 3, f"HITOS CLAVE - ESTRATEGIA {strategy} ({effort} SEMANAS)")
    _header_row(ws, 4, ["Hito", "Semana", "Descripcion", "", "", ""])
    for i, m in enumerate(milestones):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        _data_row(ws, 5 + i, [m.get("milestone", ""), f"Semana {m.get('week', '')}", "", "", "", ""], bg)

    phases = ms.get("migration_phases", []) if ms else []
    row = 5 + len(milestones) + 2
    _section(ws, row, "FASES DE MIGRACION")
    _header_row(ws, row + 1, ["Fase", "Semanas", "Acciones principales", "", "", ""])
    for i, p in enumerate(phases):
        bg = C_GRAY if i % 2 == 0 else C_WHITE
        actions = " - ".join(p.get("actions", []))
        _data_row(ws, row + 2 + i, [p.get("name", ""), p.get("weeks", ""), actions, "", "", ""], bg)

    risks = pe.get("risks", []) if pe else []
    row2 = row + 2 + len(phases) + 2
    _section(ws, row2, "GESTION DE RIESGOS")
    _header_row(ws, row2 + 1, ["Riesgo", "Probabilidad", "Impacto - Mitigacion", "", "", ""])
    for i, risk in enumerate(risks):
        bg = C_LIGHT_RED if risk.get("probability") == "HIGH" else (C_LIGHT_ORANGE if risk.get("probability") == "MEDIUM" else C_WHITE)
        mit = f"{risk.get('impact', '')} - {risk.get('mitigation', '')}"
        _data_row(ws, row2 + 2 + i, [risk.get("risk", ""), risk.get("probability", ""), mit, "", "", ""], bg)

    row3 = row2 + 2 + len(risks) + 2
    _section(ws, row3, "RIESGO DE NO ACTUAR")
    _header_row(ws, row3 + 1, ["Concepto", "Estimacion", "", "", "", ""])
    inaction = [
        ("Multas regulatorias PCI-DSS / SOX", "USD 10,000-60,000 / ano"),
        ("Exposicion GDPR",                   "Hasta 4% de facturacion global"),
        ("Crecimiento deuda tecnica",          "~25% anual"),
        ("Costo total de inaccion a 3 anos",   "USD 936,000 - 4,236,000"),
    ]
    for i, (concept, est) in enumerate(inaction):
        bg = C_LIGHT_RED if i == len(inaction) - 1 else (C_GRAY if i % 2 == 0 else C_WHITE)
        _data_row(ws, row3 + 2 + i, [concept, est, "", "", "", ""], bg)

# ── Main ──────────────────────────────────────────────────────────────────────

def generar_excel(bucket=BUCKET, prefix=PREFIX, out_dir=OUT_DIR):
    print(f"\n[EXCEL REPORT] Generando reporte ejecutivo...")
    r, c, m, dq, ms, pe, bc = _load_data(bucket, prefix)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _hoja_resumen(wb, r, c, m, dq)
    _hoja_scores(wb, r, c, m)
    _hoja_costos(wb, bc, pe)
    _hoja_equipo(wb, pe, ms)
    _hoja_roadmap(wb, ms, pe)

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "executive_report.xlsx")
    wb.save(out_path)
    print(f"  OK: {out_path}")
    return out_path


def main():
    p = argparse.ArgumentParser(description="Genera reporte ejecutivo Excel")
    p.add_argument("--bucket", default=BUCKET)
    p.add_argument("--prefix", default=PREFIX)
    a = p.parse_args()
    generar_excel(a.bucket, a.prefix)


if __name__ == "__main__":
    main()
